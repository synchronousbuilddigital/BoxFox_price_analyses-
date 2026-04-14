"""
fetch_pricing.py
Creates pricing_calculator.xlsx - a fully working local Excel file where:
  - Sheet2 = input form (change Category, Sub Category, Qty here)
  - Main   = pricing output with Excel-compatible formulas
  - Price, Product sheets = copied as values (lookup tables)
"""

import os
import gspread
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.oauth2.service_account import Credentials

BASE        = os.path.dirname(__file__)
CREDS_FILE  = os.path.join(BASE, "../credentials.json")
OUTPUT_FILE = os.path.join(BASE, "../data/pricing_calculator.xlsx")
SPREADSHEET = "1AsYmuTr6KkFtiBArIv4BTk4XB80sbOwJRXY-mQkkFuY"
SCOPES      = ['https://www.googleapis.com/auth/spreadsheets.readonly']

# Google Sheets functions not in Excel -> convert to Excel equivalents
GS_TO_EXCEL = {
    'IFNA(':   'IFERROR(',
    'IFS(':    'IFS(',       # Excel 2019+ supports IFS
    'UNIQUE(': 'UNIQUE(',    # Excel 365 supports UNIQUE
    'ROUNDUP(': 'ROUNDUP(',
}


def ifs_to_nested_if(ifs_args_str):
    """
    Convert IFS(c1,v1,c2,v2,...) arguments string to nested IF(c1,v1,IF(c2,v2,...,0))
    ifs_args_str is everything inside IFS(...) — we parse pairs carefully.
    """
    # Split by comma but respect nested parentheses
    parts = []
    depth = 0
    current = []
    for ch in ifs_args_str:
        if ch == ',' and depth == 0:
            parts.append(''.join(current).strip())
            current = []
        else:
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
            current.append(ch)
    if current:
        parts.append(''.join(current).strip())

    # Build nested IF from pairs
    if len(parts) < 2:
        return ifs_args_str

    pairs = [(parts[i], parts[i+1]) for i in range(0, len(parts)-1, 2)]
    result = '0'
    for cond, val in reversed(pairs):
        result = f'IF({cond},{val},{result})'
    return result


def fix_formula(f):
    """Convert Google Sheets formula to Excel-compatible formula."""
    if not isinstance(f, str) or not f.startswith('='):
        return f

    # Drop Google Sheets-only functions entirely
    if 'SPLIT(' in f or 'QUERY(' in f or 'UNIQUE(' in f or 'SORT(' in f:
        return ''

    # IFNA -> IFERROR with fallback ""
    f = f.replace('IFNA(', 'IFERROR(')

    # Convert IFS(...) to nested IF(...)
    # Find all IFS( occurrences and replace
    import re

    def replace_ifs(match):
        inner = match.group(1)
        return ifs_to_nested_if(inner)

    # Match IFS( ... ) — handle nested parens
    result = []
    i = 0
    while i < len(f):
        if f[i:i+4] == 'IFS(':
            # Find matching closing paren
            depth = 1
            j = i + 4
            while j < len(f) and depth > 0:
                if f[j] == '(':
                    depth += 1
                elif f[j] == ')':
                    depth -= 1
                j += 1
            inner = f[i+4:j-1]
            result.append(ifs_to_nested_if(inner))
            i = j
        else:
            result.append(f[i])
            i += 1

    return ''.join(result)


def copy_as_values(gc_sheet, sheet_name, wb):
    """Copy a Google Sheet as plain values (no formulas) - for lookup tables."""
    print(f"  Copying {sheet_name} (values)...")
    try:
        gs_ws = gc_sheet.worksheet(sheet_name)
        data  = gs_ws.get_all_values(value_render_option='FORMATTED_VALUE')
    except Exception as e:
        print(f"    Skipped: {e}")
        return

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    for r_idx, row in enumerate(data, 1):
        for c_idx, val in enumerate(row, 1):
            if val == '' or val is None:
                continue
            cell = ws.cell(row=r_idx, column=c_idx)
            # Try numeric conversion
            try:
                v = val.replace(',', '').replace('₹', '').strip()
                if '.' in v:
                    cell.value = float(v)
                else:
                    cell.value = int(v)
            except (ValueError, AttributeError):
                cell.value = val

    print(f"    {ws.max_row} rows x {ws.max_column} cols")


def copy_with_formulas(gc_sheet, sheet_name, wb):
    """Copy a Google Sheet preserving Excel-compatible formulas."""
    print(f"  Copying {sheet_name} (formulas)...")
    try:
        gs_ws    = gc_sheet.worksheet(sheet_name)
        formulas = gs_ws.get_all_values(value_render_option='FORMULA')
        values   = gs_ws.get_all_values(value_render_option='FORMATTED_VALUE')
    except Exception as e:
        print(f"    Skipped: {e}")
        return

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    for r_idx, (frow, vrow) in enumerate(zip(formulas, values), 1):
        for c_idx, (f, v) in enumerate(zip(frow, vrow), 1):
            cell = ws.cell(row=r_idx, column=c_idx)

            if isinstance(f, str) and f.startswith('='):
                fixed = fix_formula(f)
                if fixed and fixed.startswith('='):
                    cell.value = fixed
                else:
                    # Fall back to computed value
                    try:
                        nv = v.replace(',', '').replace('₹', '').strip()
                        if '.' in nv:
                            cell.value = float(nv)
                        else:
                            cell.value = int(nv)
                    except (ValueError, AttributeError):
                        cell.value = v if v != '' else None
            elif f == '' or f is None:
                cell.value = None
            else:
                try:
                    sv = str(f).replace(',', '').replace('₹', '').strip()
                    if '.' in sv:
                        cell.value = float(sv)
                    else:
                        cell.value = int(sv)
                except (ValueError, TypeError):
                    cell.value = f

    print(f"    {ws.max_row} rows x {ws.max_column} cols")


def build_input_sheet(wb, categories_map):
    """Build a clean Sheet2 input form with dropdowns."""
    if 'Sheet2' in wb.sheetnames:
        del wb['Sheet2']

    ws = wb.create_sheet(title='Sheet2', index=0)

    # Styles
    header_fill  = PatternFill("solid", fgColor="4F46E5")
    header_font  = Font(color="FFFFFF", bold=True)
    input_fill   = PatternFill("solid", fgColor="EEF2FF")
    label_font   = Font(bold=True)
    center       = Alignment(horizontal='center', vertical='center')
    thin         = Side(style='thin', color='D1D5DB')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells('A1:P1')
    title_cell = ws['A1']
    title_cell.value = "PRICING CALCULATOR - INPUT FORM"
    title_cell.font  = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill  = PatternFill("solid", fgColor="1E1B4B")
    title_cell.alignment = center

    # Instructions
    ws.merge_cells('A2:P2')
    ws['A2'].value = "Fill in the yellow cells below. Prices will calculate automatically in the MAIN sheet."
    ws['A2'].font  = Font(italic=True, color="6B7280")
    ws['A2'].alignment = center

    # Block headers
    block_headers = ['Category', 'Sub Category', 'Specifications', 'No. of Ups',
                     'Machine', 'Sheet W', 'Sheet H', 'GSM',
                     'Final Qty', 'Designing', 'Leafing/Embossing', 'Window',
                     'Pasting', 'Double Charges', 'Die Cutting', 'Notes']

    # Write 4 product blocks (rows 4,8,12,16 = data rows; 3,7,11,15 = headers)
    block_row_starts = [4, 8, 12, 16]
    block_labels     = ['Product 1', 'Product 2', 'Product 3', 'Product 4']

    # Column widths
    col_widths = [18, 18, 35, 12, 12, 10, 10, 10, 12, 12, 18, 10, 10, 14, 14, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for block_idx, (start_row, label) in enumerate(zip(block_row_starts, block_labels)):
        header_row = start_row - 1

        # Block label
        ws.merge_cells(f'A{header_row}:P{header_row}')
        lc = ws[f'A{header_row}']
        lc.value     = label
        lc.font      = header_font
        lc.fill      = header_fill
        lc.alignment = center

        # Column headers (separate row below the block label)
        col_header_row = header_row + 1
        ws.row_dimensions[col_header_row].height = 18
        for c_idx, h in enumerate(block_headers, 1):
            cell = ws.cell(row=col_header_row, column=c_idx)
            cell.value     = h
            cell.font      = label_font
            cell.fill      = PatternFill("solid", fgColor="E0E7FF")
            cell.alignment = center
            cell.border    = border

        # Input row - yellow cells (now one row below col headers)
        input_row = start_row + 1
        ws.row_dimensions[input_row].height = 22

        # Get sorted categories for this block
        cats = sorted(categories_map.keys())

        for c_idx in range(1, 17):
            cell = ws.cell(row=input_row, column=c_idx)
            cell.fill   = PatternFill("solid", fgColor="FEFCE8")
            cell.border = border
            cell.alignment = Alignment(vertical='center')

        # Pre-fill with sample data for block 1
        if block_idx == 0:
            ws.cell(row=input_row, column=1).value  = 'Bakery'
            ws.cell(row=input_row, column=2).value  = 'Brownie 1'
            ws.cell(row=input_row, column=3).value  = '89*89*38 mm | 3.5*3.5*1.5 inch'
            ws.cell(row=input_row, column=5).value  = 2029
            ws.cell(row=input_row, column=8).value  = 330
            ws.cell(row=input_row, column=9).value  = 6000
            ws.cell(row=input_row, column=10).value = 100

        # Empty row between blocks
        ws.row_dimensions[start_row + 1].height = 6

    # Category list on the side for reference
    ws['R3'].value = "AVAILABLE CATEGORIES"
    ws['R3'].font  = Font(bold=True)
    for i, cat in enumerate(sorted(categories_map.keys()), 4):
        ws.cell(row=i, column=18).value = cat

    print(f"  Built Sheet2 input form")


def build_main_pricing_sheet(gc_sheet, wb):
    """Copy Main sheet with formulas fixed for Excel."""
    print(f"  Building Main pricing sheet...")
    gs_ws    = gc_sheet.worksheet('Main')
    formulas = gs_ws.get_all_values(value_render_option='FORMULA')
    values   = gs_ws.get_all_values(value_render_option='FORMATTED_VALUE')

    if 'Main' in wb.sheetnames:
        del wb['Main']
    ws = wb.create_sheet(title='Main')

    header_fill = PatternFill("solid", fgColor="1E1B4B")
    header_font = Font(color="FFFFFF", bold=True)
    price_fill  = PatternFill("solid", fgColor="F0FDF4")

    for r_idx, (frow, vrow) in enumerate(zip(formulas, values), 1):
        for c_idx, (f, v) in enumerate(zip(frow, vrow), 1):
            cell = ws.cell(row=r_idx, column=c_idx)

            if r_idx == 1:
                # Header row
                cell.value     = f
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal='center')
                continue

            if isinstance(f, str) and f.startswith('='):
                fixed = fix_formula(f)
                if fixed and fixed.startswith('='):
                    cell.value = fixed
                else:
                    # Use computed value as fallback
                    try:
                        nv = v.replace(',', '').replace('₹', '').strip()
                        cell.value = float(nv) if '.' in nv else int(nv)
                    except (ValueError, AttributeError):
                        cell.value = v if v not in ('', '#VALUE!', '#DIV/0!', '#N/A', '#REF!') else None
            elif f == '' or f is None:
                cell.value = None
            else:
                try:
                    sv = str(f).replace(',', '').replace('₹', '').strip()
                    cell.value = float(sv) if '.' in sv else int(sv)
                except (ValueError, TypeError):
                    cell.value = f

            # Highlight price columns (col 33+)
            if c_idx >= 33 and r_idx > 1:
                cell.fill = price_fill

    # Set column widths
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 8), 30)

    print(f"  Main sheet: {ws.max_row} rows x {ws.max_column} cols")


def main():
    print("Connecting to Google Sheets...")
    creds = Credentials.from_service_account_file(CREDS_FILE, scopes=SCOPES)
    gc    = gspread.authorize(creds)
    sh    = gc.open_by_key(SPREADSHEET)
    print(f"Connected: {sh.title}\n")

    # Get categories from ALLDATA for the input form
    alldata = sh.worksheet('ALLDATA').get_all_values()
    categories_map = {}
    for row in alldata[1:]:
        cat = row[1].strip() if len(row) > 1 else ''
        sub = row[2].strip() if len(row) > 2 else ''
        if cat and cat.lower() not in ('none', 'nan', 'category'):
            if cat not in categories_map:
                categories_map[cat] = set()
            if sub:
                categories_map[cat].add(sub)

    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    print("Step 1: Building input form (Sheet2)...")
    build_input_sheet(wb, categories_map)

    print("\nStep 2: Copying lookup tables as values...")
    for sheet in ['Product', 'Product2', 'Product3', 'Product4',
                  'Product5', 'Product6', 'Product7', 'Product8', 'Price', 'Item']:
        copy_as_values(sh, sheet, wb)

    print("\nStep 3: Building Main pricing sheet...")
    build_main_pricing_sheet(sh, wb)

    # Set Sheet2 as active
    wb.active = wb['Sheet2']

    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")
    print(f"Sheets: {wb.sheetnames}")
    print()
    print("HOW TO USE:")
    print("  1. Open data/pricing_calculator.xlsx")
    print("  2. In Sheet2 - fill in Category, Sub Category, Qty (yellow cells)")
    print("  3. Go to Main sheet - all prices calculate automatically")
    print("  4. Re-run this script to sync latest data from Google Sheets")


if __name__ == "__main__":
    main()
