"""
pricing_calculator.py
=====================
Python pricing calculator that replicates the exact formulas from the
Google Sheets 'Main' tab. Takes Category + Sub Category + Qty as input
and outputs a clean Excel with all pricing options.

Usage:
    python3 scripts/pricing_calculator.py

Or import and call calculate() directly.
"""

import os
import math
import gspread
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.oauth2.service_account import Credentials
from collections import defaultdict

BASE       = os.path.dirname(__file__)
CREDS_FILE = os.path.join(BASE, "../credentials.json")
OUTPUT     = os.path.join(BASE, "../data/pricing_output.xlsx")
SHEET_KEY  = "1AsYmuTr6KkFtiBArIv4BTk4XB80sbOwJRXY-mQkkFuY"
SCOPES     = ['https://www.googleapis.com/auth/spreadsheets.readonly']


# ─────────────────────────────────────────────────────────────────────────────
# 1. DATA LOADER
# ─────────────────────────────────────────────────────────────────────────────

def load_all_data():
    """Fetch all required sheets from Google Sheets."""
    print("Connecting to Google Sheets...")
    creds = Credentials.from_service_account_file(CREDS_FILE, scopes=SCOPES)
    gc    = gspread.authorize(creds)
    sh    = gc.open_by_key(SHEET_KEY)
    print(f"Connected: {sh.title}\n")

    def get_sheet(name):
        ws   = sh.worksheet(name)
        rows = ws.get_all_values(value_render_option='FORMATTED_VALUE')
        return rows

    data = {}
    for name in ['ALLDATA', 'Price', 'Product']:
        print(f"  Loading {name}...")
        data[name] = get_sheet(name)

    return data


def parse_num(val, default=0):
    """Safely parse a number from string."""
    if val is None:
        return default
    try:
        v = str(val).replace(',', '').replace('₹', '').replace('%', '').strip()
        if v in ('', '#DIV/0!', '#VALUE!', '#N/A', '#REF!', '#NAME?', 'N/A'):
            return default
        return float(v)
    except (ValueError, TypeError):
        return default


# ─────────────────────────────────────────────────────────────────────────────
# 2. PRICE SHEET RATES  (exact cell references from Main sheet formulas)
# ─────────────────────────────────────────────────────────────────────────────

def extract_rates(price_rows):
    """Extract all rate constants from the Price sheet."""
    def p(r, c):
        return parse_num(price_rows[r-1][c-1] if len(price_rows) >= r and len(price_rows[r-1]) >= c else 0)

    rates = {
        # Lamination / coating rates (per sq cm per sheet)
        'lam_thermal':      p(2, 20),   # T2
        'lam_gloss':        p(2, 19),   # S2
        'lam_matt':         p(3, 19),   # S3
        'varnish':          p(2, 22),   # V2
        'uv_flat':          p(2, 23),   # W2
        'uv_hybrid':        p(2, 21),   # U2
        'uv_crystal':       p(7, 25),   # Y7
        'min_uv_charges':   p(2, 25),   # Y2  = 1100
        'gumming_full':     p(7, 18),   # R7
        'gumming_tb':       p(7, 19),   # S7  (top/bottom)
        'dangler_rivet':    p(7, 21),   # U7  = 50
        'dangler_elastic':  p(7, 22),   # V7  = 100
        'carry_single':     p(12, 21),  # U12 = 5
        'carry_double':     p(12, 22),  # V12 = 6
        'blister_gumming':  p(2, 26),   # Z2
        'half_cut':         p(7, 23),   # W7
        'cutting_rate':     p(7, 24),   # X7
    }

    # Printing lookup table: sheet_qty -> (print_1926, die, print_2029)
    # Cols AA(27), AB(28), AC(29=print1926), AD(30=die), AF(32=print2029)
    printing_table = []
    for row in price_rows[1:]:
        aa = parse_num(row[26] if len(row) > 26 else '')
        ab = parse_num(row[27] if len(row) > 27 else '')
        ac = parse_num(row[28] if len(row) > 28 else '')
        ad = parse_num(row[29] if len(row) > 29 else '')
        af = parse_num(row[31] if len(row) > 31 else '')
        if aa > 0:
            printing_table.append({
                'qty_from':    aa,
                'qty_to':      ab,
                'print_1926':  ac,
                'die':         ad,
                'print_2029':  af,
            })

    rates['printing_table'] = printing_table
    return rates


def lookup_printing(sheet_qty, machine, rates):
    """LOOKUP(sheet_qty, qty_ranges, printing_cost) — matches Main sheet formula."""
    table = rates['printing_table']
    result_print = 0
    result_die   = 0
    for row in table:
        if sheet_qty >= row['qty_from']:
            result_print = row['print_2029'] if machine == 2029 else row['print_1926']
            result_die   = row['die']
    return result_print, result_die


# ─────────────────────────────────────────────────────────────────────────────
# 3. PRODUCT LOOKUP  (get specs for a category+sub+spec from ALLDATA)
# ─────────────────────────────────────────────────────────────────────────────

def build_product_index(alldata_rows):
    """Index ALLDATA by (category, sub_category, specifications)."""
    index = {}
    header = alldata_rows[0]
    col = {h.strip(): i for i, h in enumerate(header)}

    for row in alldata_rows[1:]:
        cat  = row[col.get('Category', 1)].strip()       if len(row) > 1 else ''
        sub  = row[col.get('Sub Category', 2)].strip()   if len(row) > 2 else ''
        spec = row[col.get('Specifications', 3)].strip() if len(row) > 3 else ''
        if not cat:
            continue

        key = (cat.lower(), sub.lower(), spec.lower())
        index[key] = {
            'category':    cat,
            'sub_cat':     sub,
            'spec':        spec,
            'ups':         parse_num(row[col.get('No.  of Ups', 4)] if len(row) > 4 else 0),
            'machine':     parse_num(row[col.get('Machine', 5)]     if len(row) > 5 else 0),
            'sheet_w':     parse_num(row[col.get('Sheet W', 6)]     if len(row) > 6 else 0),
            'sheet_h':     parse_num(row[col.get('Sheet H', 7)]     if len(row) > 7 else 0),
            'designing':   parse_num(row[col.get('Designing', 8)]   if len(row) > 8 else 100),
            'leafing':     parse_num(row[col.get('Leafing / Embossing', 10)] if len(row) > 10 else 0),
            'window':      parse_num(row[col.get('Window', 11)]     if len(row) > 11 else 0),
            'pasting':     parse_num(row[col.get('Pasting', 12)]    if len(row) > 12 else 0),
            'double_chg':  parse_num(row[col.get('Double Charges', 13)] if len(row) > 13 else 1),
        }

    return index


def get_product(index, category, sub_cat, spec=''):
    """Find product specs, trying progressively looser matches."""
    # Exact match
    key = (category.lower(), sub_cat.lower(), spec.lower())
    if key in index:
        return index[key]
    # Without spec
    for k, v in index.items():
        if k[0] == category.lower() and k[1] == sub_cat.lower():
            return v
    # Just category
    for k, v in index.items():
        if k[0] == category.lower():
            return v
    return None


def get_all_specs(index, category, sub_cat):
    """Get all specifications for a category+sub_cat."""
    results = []
    for k, v in index.items():
        if k[0] == category.lower() and k[1] == sub_cat.lower():
            results.append(v)
    return results


# ─────────────────────────────────────────────────────────────────────────────
# 4. CORE PRICING ENGINE  (replicates Main sheet formulas exactly)
# ─────────────────────────────────────────────────────────────────────────────

# SBS paper cost per GSM (from Product sheet col CC rows)
SBS_GSM_RATE = {
    230: 85, 250: 80, 280: 82, 300: None, 330: None
}

# Default GSM per material
MATERIAL_DEFAULT_GSM = {
    'SBS':       280,
    'WhiteBack': 280,
    'GreyBack':  280,
    'Art Card':  300,
    'Maplitho':  90,
}

def roundup(x):
    return math.ceil(x)


def calc_sheet_cost(sheet_w, sheet_h, sheet_qty, gsm, material_rate):
    """
    SBS/WhiteBack/GreyBack sheet cost formula:
    =ROUNDUP((((sheet_w * sheet_h)/1550) * (gsm/1000)) * (material_rate+2) * sheet_qty)
       + ((sheet_qty/144)*15)
    """
    area_factor  = (sheet_w * sheet_h) / 1550
    weight       = gsm / 1000
    cost_per_sht = area_factor * weight * (material_rate + 2)
    total_paper  = roundup(cost_per_sht * sheet_qty + (sheet_qty / 144) * 15)
    return total_paper


def calc_price(product, final_qty, gsm, material, rates,
               print_color='Four Colour', lamination='Lamination Thermal',
               addon='Plain', die_cutting='Yes'):
    """
    Replicate the Main sheet pricing formulas exactly.
    Returns a dict with all price components and totals.
    """
    ups       = max(product['ups'], 1)
    machine   = int(product['machine']) if product['machine'] else 2029
    sheet_w   = product['sheet_w']
    sheet_h   = product['sheet_h']
    designing = product['designing'] or 100
    pasting   = product['pasting']   or 0
    leafing   = product['leafing']   or 0
    window    = product['window']    or 0

    # Sheet Qty = ROUNDUP(final_qty / ups) + 80
    sheet_qty = roundup(final_qty / ups) + 80

    # Paper cost per sheet based on material
    mat_rate_map = {
        'SBS':       85,   # per 1000 sheets at 280gsm (from Price!P2)
        'WhiteBack': 78,   # Price!Q2
        'GreyBack':  70,   # Price!R2
        'Art Card':  115,  # Price!P7
        'Maplitho':  78,   # Price!Q7
    }
    mat_rate = mat_rate_map.get(material, 82)

    # Paper / sheet cost
    paper_cost = calc_sheet_cost(sheet_w, sheet_h, sheet_qty, gsm, mat_rate)

    # Printing cost (lookup by sheet_qty and machine)
    print_cost, die_cost = lookup_printing(sheet_qty, machine, rates)
    # Multiply by printing colour factor (from Product!BB15 = 4 for four colour)
    colour_factor_map = {
        'Single Colour':      1,
        'Double Colour':      2,
        'Four Colour':        4,
        'Four + One Colour':  6.75,
        'Four + Two Colour':  7.75,
        'Four + Four Colour': 9.75,
        'Without Print':      0,
    }
    colour_factor = colour_factor_map.get(print_color, 4)
    print_cost = roundup(print_cost * colour_factor)

    # Die cost
    die_cost = die_cost if die_cutting == 'Yes' else 0

    # Designing
    designing_cost = designing

    # Lamination / coating cost per unit
    # Formula: MAX((sheet_w * sheet_h * rate * sheet_qty) / final_qty, 300/final_qty)
    lam_rate_map = {
        'Lamination Thermal':      rates['lam_thermal'],
        'Lamination Normal Gloss': rates['lam_gloss'],
        'Lamination Normal Matt':  rates['lam_matt'],
        'Varnish':                 rates['varnish'],
        'UV Flat':                 rates['uv_flat'],
        'UV Hybrid':               rates['uv_hybrid'],
        'UV Crystal':              rates['uv_crystal'],
        'Spot UV':                 rates['uv_flat'],
        'Plain':                   0,
    }
    lam_rate = lam_rate_map.get(lamination, rates['lam_thermal'])

    if lam_rate > 0:
        if lamination in ('UV Flat', 'UV Hybrid', 'UV Crystal', 'Spot UV'):
            lam_cost = max(
                ((sheet_w * sheet_h * lam_rate * sheet_qty) / final_qty) + (350 / final_qty),
                500 / final_qty
            )
            if lamination == 'UV Crystal':
                lam_cost += 300 / final_qty  # AW2 = blister gumming
        else:
            lam_cost = max(
                (sheet_w * sheet_h * lam_rate * sheet_qty) / final_qty,
                300 / final_qty
            )
    else:
        lam_cost = 0

    # Add-on cost per unit
    addon_rate_map = {
        'Carry Bag Single Pasting': rates['carry_single'],
        'Carry Bag Double Pasting': rates['carry_double'],
        'Gumming Full':             rates['gumming_full'],
        'Gumming Top Bottom':       rates['gumming_tb'],
        'Dangler Making':           0,
        'Dangler With Rivet':       rates['dangler_rivet'] / final_qty,
        'Dangler With Rivet and Elastic': rates['dangler_elastic'] / final_qty,
        'Plain':                    0,
        'Cutting':                  0,
    }
    addon_cost = addon_rate_map.get(addon, 0)

    # Other charges = ROUNDUP((final_qty * pasting) + (final_qty * pasting * 5%))
    #                 + MAX((sheet_qty * ups * addon_cost), 200) + 1
    other_charges = roundup(
        (final_qty * pasting) + (final_qty * pasting * 0.05) +
        max((sheet_qty * ups * addon_cost), 200) + 1
    )

    # Total (base) = ROUNDUP(paper + print + die + designing + other_charges)
    total_base = roundup(paper_cost + print_cost + die_cost + designing_cost + other_charges)

    # Per unit prices for each material option
    # Formula: (paper + designing + print + other) / final_qty
    # Note: die_cost is a one-time job cost, not per-unit in the per-unit columns
    def per_unit(paper, extra=0):
        t = paper + designing_cost + print_cost + other_charges
        return round((t + extra) / final_qty, 4)

    sbs_paper       = calc_sheet_cost(sheet_w, sheet_h, sheet_qty, gsm, 85)
    wb_paper        = calc_sheet_cost(sheet_w, sheet_h, sheet_qty, gsm, 78)
    gb_paper        = calc_sheet_cost(sheet_w, sheet_h, sheet_qty, gsm, 70)
    ac_paper        = calc_sheet_cost(sheet_w, sheet_h, sheet_qty, gsm, 115)
    mp_paper        = calc_sheet_cost(sheet_w, sheet_h, sheet_qty, gsm, 78)

    return {
        # Inputs
        'category':        product['category'],
        'sub_category':    product['sub_cat'],
        'specifications':  product['spec'],
        'final_qty':       final_qty,
        'sheet_qty':       sheet_qty,
        'ups':             ups,
        'machine':         machine,
        'sheet_w':         sheet_w,
        'sheet_h':         sheet_h,
        'gsm':             gsm,
        'material':        material,
        'print_color':     print_color,
        'lamination':      lamination,
        'add_on':          addon,
        'die_cutting':     die_cutting,

        # Cost components
        'paper_cost':      paper_cost,
        'print_cost':      print_cost,
        'die_cost':        die_cost,
        'designing':       designing_cost,
        'other_charges':   other_charges,
        'total_cost':      total_base,

        # Per-unit price by material
        'price_per_unit_SBS':       per_unit(sbs_paper),
        'price_per_unit_WhiteBack': per_unit(wb_paper),
        'price_per_unit_GreyBack':  per_unit(gb_paper),
        'price_per_unit_ArtCard':   per_unit(ac_paper),
        'price_per_unit_Maplitho':  per_unit(mp_paper),

        # Lamination add-on per unit
        'lam_cost_per_unit': round(lam_cost, 4),

        # Final price per unit (selected material + lamination)
        'final_price_per_unit': round(per_unit(paper_cost) + lam_cost, 4),
        'final_total':          roundup((per_unit(paper_cost) + lam_cost) * final_qty),
    }


# ─────────────────────────────────────────────────────────────────────────────
# 5. EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

def write_excel(results, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pricing Output"

    # Styles
    h_fill  = PatternFill("solid", fgColor="1E1B4B")
    h_font  = Font(color="FFFFFF", bold=True, size=10)
    s_fill  = PatternFill("solid", fgColor="EEF2FF")
    p_fill  = PatternFill("solid", fgColor="F0FDF4")
    t_fill  = PatternFill("solid", fgColor="FEF9C3")
    center  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left    = Alignment(horizontal='left',   vertical='center')
    thin    = Side(style='thin', color='D1D5DB')
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells('A1:V1')
    ws['A1'].value     = "PRICING CALCULATOR OUTPUT"
    ws['A1'].font      = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill      = PatternFill("solid", fgColor="0F0A3C")
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    # Headers
    headers = [
        'Category', 'Sub Category', 'Specifications', 'Final Qty', 'Sheet Qty',
        'Ups', 'Machine', 'Sheet W', 'Sheet H', 'GSM',
        'Material', 'Print Color', 'Lamination', 'Add On', 'Die Cutting',
        'Paper Cost', 'Print Cost', 'Die Cost', 'Designing', 'Other Charges',
        'Total Cost (₹)',
        'Price/Unit SBS', 'Price/Unit WhiteBack', 'Price/Unit GreyBack',
        'Price/Unit ArtCard', 'Price/Unit Maplitho',
        'Lam Cost/Unit', 'Final Price/Unit', 'Final Total (₹)'
    ]

    col_widths = [14, 16, 35, 10, 10, 6, 10, 9, 9, 8,
                  12, 16, 22, 22, 12,
                  12, 12, 10, 12, 14,
                  14,
                  15, 18, 17, 16, 17,
                  14, 16, 14]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=c)
        cell.value     = h
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.row_dimensions[2].height = 36

    # Data rows
    keys = [
        'category', 'sub_category', 'specifications', 'final_qty', 'sheet_qty',
        'ups', 'machine', 'sheet_w', 'sheet_h', 'gsm',
        'material', 'print_color', 'lamination', 'add_on', 'die_cutting',
        'paper_cost', 'print_cost', 'die_cost', 'designing', 'other_charges',
        'total_cost',
        'price_per_unit_SBS', 'price_per_unit_WhiteBack', 'price_per_unit_GreyBack',
        'price_per_unit_ArtCard', 'price_per_unit_Maplitho',
        'lam_cost_per_unit', 'final_price_per_unit', 'final_total'
    ]

    for r_idx, res in enumerate(results, 3):
        fill = s_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for c_idx, key in enumerate(keys, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value  = res.get(key, '')
            cell.border = border
            cell.alignment = left

            # Highlight price columns
            if c_idx >= 16 and c_idx <= 21:
                cell.fill = p_fill
            elif c_idx >= 22:
                cell.fill = t_fill
            else:
                cell.fill = fill

            # Format numbers
            if isinstance(cell.value, float):
                if c_idx in (22, 23, 24, 25, 26, 27, 28):
                    cell.number_format = '₹#,##0.0000'
                elif c_idx in (16, 17, 18, 19, 20, 21, 29):
                    cell.number_format = '₹#,##0'

        ws.row_dimensions[r_idx].height = 18

    # Freeze header rows
    ws.freeze_panes = 'A3'

    wb.save(output_path)
    print(f"\nSaved: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# 6. MAIN — define your inputs here
# ─────────────────────────────────────────────────────────────────────────────

JOBS = [
    # (Category,    Sub Category,  Specifications,                      Qty,   GSM, Material,    Print Color,          Lamination,              AddOn,                    Die)
    ('Bakery',      'Brownie 1',   '89*89*38 mm | 3.5*3.5*1.5 inch',   6000,  330, 'SBS',       'Four Colour',        'Lamination Thermal',    'Plain',                  'Yes'),
    ('Bakery',      'Brownie 1',   '89*89*38 mm | 3.5*3.5*1.5 inch',   2000,  330, 'SBS',       'Four Colour',        'Lamination Thermal',    'Plain',                  'Yes'),
    ('Bakery',      'Brownie 1',   '89*89*38 mm | 3.5*3.5*1.5 inch',   1000,  330, 'SBS',       'Four Colour',        'Lamination Thermal',    'Plain',                  'Yes'),
    ('Carry Bag',   'Cashify',     '240*64*160 mm | 9.5*2.5*6.25 inch', 20000, 170, 'WhiteBack', 'Four Colour',        'Lamination Thermal',    'Carry Bag Single Pasting','Yes'),
    ('Carry Bag',   'Cashify',     '240*64*160 mm | 9.5*2.5*6.25 inch', 35000, 170, 'WhiteBack', 'Four Colour',        'Lamination Thermal',    'Carry Bag Single Pasting','Yes'),
    ('Flyer',       'A5',          '148 × 210 millimeters',             100000, 90, 'Art Card',  'Four Colour',        'Lamination Normal Gloss','Plain',                 'No'),
    ('Flyer',       'A5',          '148 × 210 millimeters',             50000,  90, 'Art Card',  'Four Colour',        'Lamination Normal Gloss','Plain',                 'No'),
]


def main():
    raw   = load_all_data()
    rates = extract_rates(raw['Price'])
    index = build_product_index(raw['ALLDATA'])

    print(f"Loaded {len(index)} products from ALLDATA")
    print(f"Loaded {len(rates['printing_table'])} printing rate rows\n")

    results = []
    for cat, sub, spec, qty, gsm, material, print_color, lamination, addon, die in JOBS:
        product = get_product(index, cat, sub, spec)
        if not product:
            print(f"  WARNING: Product not found: {cat} > {sub}")
            continue

        result = calc_price(
            product     = product,
            final_qty   = qty,
            gsm         = gsm,
            material    = material,
            rates       = rates,
            print_color = print_color,
            lamination  = lamination,
            addon       = addon,
            die_cutting = die,
        )
        results.append(result)

        print(f"  {cat} > {sub} | Qty:{qty:>7,} | "
              f"Total: ₹{result['total_cost']:>8,} | "
              f"Per Unit ({material}): ₹{result['price_per_unit_SBS'] if material=='SBS' else result['price_per_unit_WhiteBack']:.4f} | "
              f"Final/Unit: ₹{result['final_price_per_unit']:.4f}")

    write_excel(results, OUTPUT)
    print(f"\nDone. Open: {OUTPUT}")


if __name__ == "__main__":
    main()
